import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os, sys, importlib.util

# ──────────────────────────────────────────
# tc_data 파일 로드
# ──────────────────────────────────────────
if len(sys.argv) < 2:
    print("Usage: python generate_tc.py tc_data/<화면명>.py")
    sys.exit(1)

tc_file = sys.argv[1]
spec = importlib.util.spec_from_file_location("tc_module", tc_file)
mod  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

screen_name = mod.screen_name
tc_data     = mod.tc_data

# ──────────────────────────────────────────
# 스타일 헬퍼
# ──────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def border_thin():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

FILL_HEADER = fill("BFBFBF")
FILL_NONE   = fill("FFFFFF")

FONT_TITLE  = Font(name="Arial", size=18, bold=True, color="002060")
FONT_HEADER = Font(name="Arial", size=9,  bold=True)
FONT_DATA   = Font(name="Arial", size=9)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ──────────────────────────────────────────
# 워크북 생성
# ──────────────────────────────────────────
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(output_dir, exist_ok=True)
file_path = os.path.join(output_dir, f"QA_{screen_name}_TC.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
tc_sheet_title = f"{screen_name} TC"
ws.title = tc_sheet_title

# ── 컬럼 너비 ──
col_widths = {
    "A": 13.75, "B": 20.25, "C": 22.13, "D": 22.75,
    "E": 24.0,  "F": 22.13, "G": 8.25,  "H": 32.38,
    "I": 12.63, "J": 10.0,  "K": 10.0,  "L": 13.63,
    "M": 43.5,
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# ──────────────────────────────────────────
# 행 1: 우상단 요약 헤더
# ──────────────────────────────────────────
for col, val in [(18, "Priority"), (19, "개수")]:
    c = ws.cell(row=1, column=col, value=val)
    c.fill = fill("D9D9D9")
    c.font = Font(name="Arial", size=9, bold=True)
    c.alignment = ALIGN_CENTER
    c.border = border_thin()

# ──────────────────────────────────────────
# 행 2: 제목 + 우상단 P1 집계
# ──────────────────────────────────────────
ws.merge_cells("A2:M2")
c = ws.cell(row=2, column=1, value=screen_name)
c.font = FONT_TITLE
c.alignment = ALIGN_LEFT

for col, val in [(18, "P1"), (19, "=COUNTIF(G:G,R2)")]:
    c = ws.cell(row=2, column=col, value=val)
    c.font = FONT_DATA
    c.alignment = ALIGN_CENTER
    c.border = border_thin()

# ──────────────────────────────────────────
# 행 3~4: 컬럼 헤더
# ──────────────────────────────────────────
single_headers = [
    ("A", "1 Depth"), ("B", "2 Depth"), ("C", "3 Depth"), ("D", "4 Depth"),
    ("E", "5 Depth"), ("F", "6 Depth"), ("G", "Priority"),
    ("H", "Expected Result"), ("M", "합계"),
]
for col_letter, val in single_headers:
    col = ord(col_letter) - ord("A") + 1
    ws.merge_cells(f"{col_letter}3:{col_letter}4")
    c = ws.cell(row=3, column=col, value=val)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = ALIGN_CENTER
    c.border = border_thin()

ws.merge_cells("I3:L3")
c = ws.cell(row=3, column=9, value="Test Result")
c.fill = FILL_HEADER
c.font = FONT_HEADER
c.alignment = ALIGN_CENTER
c.border = border_thin()

sub_headers = [(9, "Android"), (10, "iOS"), (11, "Date"), (12, "BTS ID")]
for col, val in sub_headers:
    c = ws.cell(row=4, column=col, value=val)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = ALIGN_CENTER
    c.border = border_thin()

for row, p_val, cnt_val in [(3, "P2", "=COUNTIF(G:G,R3)"), (4, "P3", "=COUNTIF(G:G,R4)")]:
    ws.cell(row=row, column=18, value=p_val).font = FONT_DATA
    ws.cell(row=row, column=19, value=cnt_val).font = FONT_DATA
    for col in (18, 19):
        ws.cell(row=row, column=col).alignment = ALIGN_CENTER
        ws.cell(row=row, column=col).border = border_thin()

ws.cell(row=3, column=17, value="=SUM(S2:S4)").font = FONT_DATA
ws.freeze_panes = "A5"

# ──────────────────────────────────────────
# 데이터 행 (행 5부터)
# ──────────────────────────────────────────
for row_idx, row in enumerate(tc_data, start=5):
    d1, d2, d3, d4, d5, d6, priority, expected = row

    for col_idx, val in enumerate([d1, d2, d3, d4, d5, d6, priority, expected], start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
        c.font = FONT_DATA
        c.alignment = ALIGN_CENTER if col_idx == 7 else ALIGN_LEFT
        c.border = border_thin()

    for col_idx in range(9, 13):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = FONT_DATA
        c.border = border_thin()

    ws.row_dimensions[row_idx].height = 30

ws.auto_filter.ref = "A4:L4"

# ──────────────────────────────────────────
# A~D열 (1~4 Depth) 연속 동일 값 셀 병합
# 상위 Depth가 같을 때만 하위 Depth 병합
# ──────────────────────────────────────────
data_start = 5
data_end   = 4 + len(tc_data)
col_letters = ["A", "B", "C", "D"]  # 1~4 Depth

for col_idx, col_letter in enumerate(col_letters, start=1):
    merge_start = data_start
    for r in range(data_start + 1, data_end + 2):
        # 현재 행과 병합 시작 행의 1~col_idx 값이 모두 같아야 병합 유지
        same_group = all(
            ws.cell(row=r, column=c).value == ws.cell(row=merge_start, column=c).value
            for c in range(1, col_idx + 1)
        )

        if not same_group or r == data_end + 1:
            if r - 1 > merge_start:
                ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{r - 1}")
                c = ws.cell(row=merge_start, column=col_idx)
                c.alignment = ALIGN_CENTER
                c.font      = FONT_DATA
                c.border    = border_thin()
            merge_start = r

# ──────────────────────────────────────────
# 데이터 유효성 검사: Android(I), iOS(J) 컬럼 → P / F / NA / NT
# ──────────────────────────────────────────
last_row = 4 + len(tc_data)
dv = DataValidation(
    type="list",
    formula1='"P,F,NA,NT"',
    allow_blank=True,
    showDropDown=False,
)
dv.sqref = f"I5:J{last_row}"
ws.add_data_validation(dv)

# ──────────────────────────────────────────
# 조건부 서식: F 값 → 연한 빨간색 배경
# ──────────────────────────────────────────
red_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
ws.conditional_formatting.add(
    f"I5:J{last_row}",
    CellIsRule(operator="equal", formula=['"F"'], fill=red_fill),
)

# ──────────────────────────────────────────
# Summary 탭 (검증 결과서)
# ──────────────────────────────────────────
summary = wb.create_sheet(title="Summary", index=0)

# 열 너비: B는 "총 합계 (단말 1개 기준 Case)" 담을 만큼 넓게, 나머지는 숫자/라벨용
summary_col_widths = {
    "A": 2.5, "B": 32, "C": 11, "D": 11, "E": 11, "F": 11,
    "G": 11, "H": 11, "I": 11, "J": 11, "K": 12, "L": 12, "M": 12,
}
for col_letter, width in summary_col_widths.items():
    summary.column_dimensions[col_letter].width = width

# 테두리 스타일
_thin  = Side(border_style="thin",   color="8F8F8F")
_med   = Side(border_style="medium", color="404040")
border_all_thin = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _box_border(r1, c1, r2, c2):
    """지정 범위 바깥쪽에 medium border를 두름 (내부는 건드리지 않음)."""
    for col in range(c1, c2 + 1):
        top = summary.cell(row=r1, column=col)
        bot = summary.cell(row=r2, column=col)
        top.border = Border(
            left=top.border.left, right=top.border.right,
            top=_med, bottom=top.border.bottom,
        )
        bot.border = Border(
            left=bot.border.left, right=bot.border.right,
            top=bot.border.top, bottom=_med,
        )
    for row in range(r1, r2 + 1):
        lft = summary.cell(row=row, column=c1)
        rgt = summary.cell(row=row, column=c2)
        lft.border = Border(
            left=_med, right=lft.border.right,
            top=lft.border.top, bottom=lft.border.bottom,
        )
        rgt.border = Border(
            left=rgt.border.left, right=_med,
            top=rgt.border.top, bottom=rgt.border.bottom,
        )

def _apply_thin_grid(r1, c1, r2, c2):
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            summary.cell(row=row, column=col).border = border_all_thin

# ── 제목 ──
summary.merge_cells("B2:M2")
c = summary.cell(row=2, column=2, value="검증 결과서")
c.font = FONT_TITLE
c.alignment = Alignment(horizontal="center", vertical="center")
summary.row_dimensions[2].height = 36

# ── 메타 정보 (범위/Test App Ver. / 결과보고일자/검증자) ──
meta_font_label = Font(name="Arial", size=10, bold=True, color="FFFFFF")
meta_font_value = Font(name="Arial", size=10)
meta_label_fill = fill("595959")
meta_value_fill = fill("FAFAFA")

meta_rows = [
    (4, "범위",          screen_name, "결과보고일자", ""),
    (5, "Test App Ver.", "",          "검증자",       ""),
]
for r, label1, val1, label2, val2 in meta_rows:
    # label1: B, value1: C:F (merged), label2: G:H (merged), value2: I:M (merged)
    summary.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    summary.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    summary.merge_cells(start_row=r, start_column=9, end_row=r, end_column=13)

    lc1 = summary.cell(row=r, column=2, value=label1)
    lc1.font = meta_font_label
    lc1.fill = meta_label_fill
    lc1.alignment = ALIGN_CENTER

    vc1 = summary.cell(row=r, column=3, value=val1)
    vc1.font = meta_font_value
    vc1.fill = meta_value_fill
    vc1.alignment = ALIGN_CENTER

    lc2 = summary.cell(row=r, column=7, value=label2)
    lc2.font = meta_font_label
    lc2.fill = meta_label_fill
    lc2.alignment = ALIGN_CENTER

    vc2 = summary.cell(row=r, column=9, value=val2)
    vc2.font = meta_font_value
    vc2.fill = meta_value_fill
    vc2.alignment = ALIGN_CENTER

    summary.row_dimensions[r].height = 24

_apply_thin_grid(4, 2, 5, 13)
_box_border(4, 2, 5, 13)

# ── 집계 테이블 헤더 ──
summary_headers = [
    "구분", "총 항목", "Not Run", "Pass", "Fail",
    "Block", "N/T", "NA", "Run", "성공율", "결함율", "수행율",
]
header_fill = fill("305496")
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
for idx, val in enumerate(summary_headers, start=2):
    c = summary.cell(row=7, column=idx, value=val)
    c.fill = header_fill
    c.font = header_font
    c.alignment = ALIGN_CENTER
summary.row_dimensions[7].height = 26

# ── 집계 수식 (TC 시트의 I열 = Android 기준) ──
last_row = 4 + len(tc_data)
sheet_ref = f"'{tc_sheet_title}'!I5:I{last_row}"
total = len(tc_data)

row_specs = [
    (8, screen_name,                         False),
    (9, "총 합계 (단말 1개 기준 Case)",        True),
]

data_font        = Font(name="Arial", size=10)
data_font_bold   = Font(name="Arial", size=10, bold=True)
total_row_fill   = fill("FFF2CC")
screen_row_fill  = fill("FFFFFF")

for r, label, is_total in row_specs:
    row_fill = total_row_fill if is_total else screen_row_fill
    row_font = data_font_bold if is_total else data_font

    cells = [
        (2, label, ALIGN_LEFT),
        (3, total, ALIGN_CENTER),
        (4, f'=$C{r}-($E{r}+$F{r}+$G{r}+$H{r}+$I{r})', ALIGN_CENTER),
        (5, f'=COUNTIF({sheet_ref},"P")', ALIGN_CENTER),
        (6, f'=COUNTIF({sheet_ref},"F")', ALIGN_CENTER),
        (7, 0, ALIGN_CENTER),
        (8, f'=COUNTIF({sheet_ref},"NT")', ALIGN_CENTER),
        (9, f'=COUNTIF({sheet_ref},"NA")', ALIGN_CENTER),
        (10, f'=$E{r}+$F{r}', ALIGN_CENTER),
        (11, f'=IF($C{r}=0,0,$E{r}/$C{r})', ALIGN_CENTER),
        (12, f'=IF($C{r}=0,0,$F{r}/$C{r})', ALIGN_CENTER),
        (13, f'=IF($C{r}=0,0,$J{r}/$C{r})', ALIGN_CENTER),
    ]
    for col, val, align in cells:
        c = summary.cell(row=r, column=col, value=val)
        c.font = row_font
        c.alignment = align
        c.fill = row_fill

    for col in (11, 12, 13):
        summary.cell(row=r, column=col).number_format = "0.00%"

    summary.row_dimensions[r].height = 24

_apply_thin_grid(7, 2, 9, 13)
_box_border(7, 2, 9, 13)

# ── 참고 및 특이사항 ──
note_header_row = 11
note_body_top   = 12
note_body_bot   = 16

summary.merge_cells(start_row=note_header_row, start_column=2,
                    end_row=note_header_row, end_column=13)
c = summary.cell(row=note_header_row, column=2, value="참고 및 특이사항")
c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
c.fill = meta_label_fill
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
summary.row_dimensions[note_header_row].height = 24

summary.merge_cells(start_row=note_body_top, start_column=2,
                    end_row=note_body_bot, end_column=13)
body = summary.cell(row=note_body_top, column=2, value="")
body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
body.fill = fill("FFFFFF")
for r in range(note_body_top, note_body_bot + 1):
    summary.row_dimensions[r].height = 20

_apply_thin_grid(note_header_row, 2, note_body_bot, 13)
_box_border(note_header_row, 2, note_body_bot, 13)

# 격자선 숨김으로 깔끔하게
summary.sheet_view.showGridLines = False

wb.save(file_path)
print(f"저장 완료: {file_path}")
print(f"총 TC: {len(tc_data)}개")
